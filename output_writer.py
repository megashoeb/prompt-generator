"""Export prompts to .txt and .xlsx formats."""

import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def remove_duplicate_style(text: str) -> str:
    """Remove duplicate style blocks from prompt text, keeping only the last one.

    Handles all four built-in style line patterns:
      • Dark Fantasy  — starts with "Hyper-detailed dark fantasy digital painting"
      • History 1     — starts with "Hand-painted oil illustration on aged"
      • Woodcut       — starts with "woodcut linocut relief print"
      • Victorian     — starts with "vintage 19th century newspaper engraving"
    """
    style_patterns = [
        r'(Hyper-detailed dark fantasy digital painting.*?16:9 aspect ratio\.?)',
        r'(Hand-painted oil illustration on aged.*?16:9 aspect ratio\.?)',
        r'(digital oil-paint realism.*?16:9 aspect ratio\.?)',           # History 2 Color
        r'(vintage monochrome charcoal sketch.*?16:9 aspect ratio\.?)',  # History 2 B&W
        r'(Classic impasto oil painting.*?16:9 aspect ratio\.?)',        # History 3
        r'(Ancient fresco on aged cracked plaster.*?16:9 aspect ratio\.?)',  # History 4
        r'(Hand-drawn 2D animation-style digital illustration.*?16:9 aspect ratio\.?)',  # History 5
        r'(woodcut linocut relief print.*?16:9)',
        r'(vintage 19th century newspaper engraving.*?16:9)',
    ]
    for style_pattern in style_patterns:
        matches = list(re.finditer(style_pattern, text, re.DOTALL | re.IGNORECASE))
        if len(matches) >= 2:
            # Remove all but the last occurrence
            for match in matches[:-1]:
                start = match.start()
                end = match.end()
                # Also eat any trailing punctuation/spaces before the next block
                while start > 0 and text[start - 1] in ' ,.\n':
                    start -= 1
                text = text[:start] + ' ' + text[end:]
            # Clean up any double spaces or ". ." artifacts left behind
            text = re.sub(r' {2,}', ' ', text)
            text = re.sub(r'\. \.', '.', text)
    return text.strip()


def count_color_bw(prompts: list[dict]) -> tuple[int, int]:
    """Count Color vs B&W prompts for History 2 style.

    Returns (color_count, bw_count).
    Detection is keyword-based on the generated prompt text.
    """
    _BW_KEYWORDS    = ('monochrome', 'charcoal sketch', 'chalk-line', 'grayscale', 'black and white')
    _COLOR_KEYWORDS = ('oil-paint realism', 'candlelight glow', 'golden highlights', 'chiaroscuro')
    color_count = bw_count = 0
    for p in prompts:
        text = p.get("image_prompt", "").lower()
        if any(kw in text for kw in _BW_KEYWORDS):
            bw_count += 1
        elif any(kw in text for kw in _COLOR_KEYWORDS):
            color_count += 1
        else:
            color_count += 1   # default to Color when ambiguous
    return color_count, bw_count


def count_noor_prompts(prompts: list[dict]) -> tuple[int, int]:
    """Count Noor (sacred figure) vs normal prompts for History 3 style.

    Returns (noor_count, normal_count).
    """
    _NOOR_KEYWORDS = ('noor light', 'divine luminance', 'blinding golden', 'obscures all physical features')
    noor_count = normal_count = 0
    for p in prompts:
        text = p.get("image_prompt", "").lower()
        if any(kw in text for kw in _NOOR_KEYWORDS):
            noor_count += 1
        else:
            normal_count += 1
    return noor_count, normal_count


def count_fire_accent_prompts(prompts: list[dict]) -> tuple[int, int]:
    """Count how many History 5 prompts include fire/torch warm glow accent.

    Returns (fire_count, no_fire_count).
    """
    _FIRE_KEYWORDS = (
        'campfire', 'torch', 'fire glow', 'ember sparks',
        'firelight', 'fire orange', 'ember red', 'warm glow spill',
    )
    fire_count = no_fire_count = 0
    for p in prompts:
        text = p.get("image_prompt", "").lower()
        if any(kw in text for kw in _FIRE_KEYWORDS):
            fire_count += 1
        else:
            no_fire_count += 1
    return fire_count, no_fire_count


def _detect_prompt_type(prompt_text: str) -> str:
    """Return 'B&W' or 'Color' for a single prompt string (History 2 use)."""
    t = prompt_text.lower()
    if any(kw in t for kw in ('monochrome', 'charcoal sketch', 'chalk-line', 'grayscale', 'black and white')):
        return "B&W"
    return "Color"


def process_prompt_with_style(scene_text: str, style_key: str,
                               subtitle_text: str = "") -> str:
    """Apply visual-style post-processing to an already-extracted scene description.

    For dark_fantasy / custom the LLM wrote the complete prompt — return unchanged.
    For woodcut / victorian_engraving:
      • Apply safety word replacements
      • Trim to word limit
      • Append scene-specific color accent + style line
    """
    try:
        from styles import STYLES, build_final_prompt
    except ImportError:
        return scene_text   # styles.py not present — no-op

    style = STYLES.get(style_key, STYLES["dark_fantasy"])
    if not style.get("append_style", False):
        return scene_text

    return build_final_prompt(scene_text, style_key, subtitle_text or scene_text)


def clean_prompt_text(text: str) -> str:
    """Clean markdown artifacts and bracket formatting from prompt text."""
    # Remove ** bold markers
    text = text.replace("**", "")

    # Remove * italic markers (standalone asterisks around words)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)

    # Remove any remaining lone * characters
    text = text.replace("*", "")

    # Remove # headers at start of lines
    text = re.sub(r'(?m)^#+\s*', '', text)

    # Remove bracket tags if they slipped through
    bracket_tags = [
        "[Subject:", "[Action:", "[Location/Context:", "[Location:",
        "[Composition:", "[Camera/Lens:", "[Camera:", "[Lens:",
        "[Color Grading:", "[Color:", "[Style:", "[Style + Lighting + Texture:",
        "[Lighting:", "[Texture:", "[Aspect Ratio:",
    ]
    for tag in bracket_tags:
        text = text.replace(tag, "")

    # Remove closing brackets
    text = text.replace("]", "")

    # Remove + separators between sections (e.g. "...detail] + [Next...")
    text = re.sub(r'\s*\+\s*(?=[A-Z])', ' ', text)

    # Remove [STYLE: ...] shorthand tags
    text = re.sub(r'\[STYLE:[^\]]*\]', '', text)

    # Remove [REFRESH ...] internal tags
    text = re.sub(r'\[REFRESH[^\]]*\]', '', text)

    # Remove character ID tags like [IBLIS-01] [TRAVELER-01]
    text = re.sub(r'\[[A-Z]+-\d+\]', '', text)

    # Clean up multiple spaces
    text = re.sub(r' {2,}', ' ', text)

    # Remove duplicate style blocks
    text = remove_duplicate_style(text)

    # ── Full mojibake fix using re-encoding trick (covers ALL variants) ────────
    def _fix_seq(m):
        """Re-encode a mojibake sequence via Latin-1 → UTF-8."""
        s = m.group(0)
        try:
            return s.encode('latin-1').decode('utf-8')
        except (UnicodeEncodeError, UnicodeDecodeError):
            return s

    # Layer 1: Ã + continuation byte — covers all Latin-1 Supplement chars
    # (à á â ã ä å æ ç è é ê ë ì í î ï ð ñ ò ó ô õ ö ø ù ú û ü ý þ ÿ + capitals)
    text = re.sub(r'Ã[\x80-\xbf]', _fix_seq, text)
    # Layer 2: Å + continuation byte — covers Latin Extended-A (ő ű ş Ş ğ Ğ ı İ etc.)
    text = re.sub(r'Å[\x80-\xbf]', _fix_seq, text)
    # Layer 3: Ä + continuation byte — covers more Latin Extended-A (ā–ŀ range)
    text = re.sub(r'Ä[\x80-\xbf]', _fix_seq, text)
    # Layer 4: â + TWO continuation bytes — em-dash, en-dash, smart quotes, etc.
    text = re.sub(r'â[\x80-\xbf][\x80-\xbf]', _fix_seq, text)
    # Layer 5: Windows-1252 variants (€=0x80 in Win-1252, not valid Latin-1)
    for _bad, _good in [
        ('â€"', '—'), ('â€"', '–'),
        ('â€™', '\u2019'), ('â€˜', '\u2018'),
        ('â€œ', '\u201c'), ('â€\x9d', '\u201d'),
        ('â€¦', '…'), ('â€¢', '•'), ('â□□', '—'), ('â□+', '—'),
    ]:
        text = text.replace(_bad, _good)
    # Layer 6: bare â — continuation bytes silently dropped
    text = text.replace('â', '-')

    # ── Expression spam cleanup ───────────────────────────────────────────────
    text = clean_expression_spam(text)

    # Clean up leading/trailing whitespace
    return text.strip()


def clean_expression_spam(text: str) -> str:
    """Remove misused 'Expression:' tags that describe non-facial content.

    'Expression:' must ONLY describe a character's facial expression or body
    language. Tags applied to scene mood, camera, palette, or abstract ideas
    are stripped — their content is kept but the 'Expression:' prefix is removed.
    """
    if 'Expression:' not in text and 'expression:' not in text:
        return text

    # Keywords that legitimately belong after 'Expression:'
    FACIAL_KEYWORDS = {
        'eyes', 'eye', 'jaw', 'mouth', 'lips', 'lip', 'brow', 'forehead',
        'gaze', 'stare', 'smile', 'frown', 'scowl', 'sneer', 'grimace',
        'wince', 'glare', 'weep', 'cry', 'laugh', 'smirk', 'snarl',
        'squint', 'determined', 'fearful', 'fierce', 'solemn', 'stoic',
        'defiant', 'terrified', 'anguish', 'serene', 'contempt', 'resolute',
        'hunched', 'posture', 'stance', 'kneel', 'cower', 'tremble', 'trembling',
        'clench', 'clenched', 'furrow', 'furrowed', 'narrow', 'narrowed',
        'wide', 'cold', 'calculating', 'thin-lipped', 'downcast', 'sovereign',
        'proud', 'haughty', 'hollow', 'vacant', 'intense', 'piercing',
        'raised eyebrow', 'open mouth', 'closed eyes', 'tear', 'tears',
        'set jaw', 'hard jaw', 'tight lips', 'parted lips', 'pressed lips',
    }

    # Tokens that indicate non-facial misuse
    NON_FACIAL_INDICATORS = {
        'palette', 'color', 'colour', 'camera', 'pan', 'zoom', 'sweep',
        'absence', 'vacuum', 'void', 'era', 'epoch', 'symboliz', 'represent',
        'fracture', 'defeat as', 'loss as', 'desaturated', 'muted tone',
        'the scene', 'the battle', 'the frame', 'the image', 'the moment',
        'the end', 'no king', 'abandoned', 'empty throne', 'scattered weapon',
    }

    def _is_facial(content: str) -> bool:
        lower = content.lower()
        if any(kw in lower for kw in NON_FACIAL_INDICATORS):
            return False
        return any(kw in lower for kw in FACIAL_KEYWORDS)

    # Split on 'Expression:' (case-insensitive), process each segment
    parts = re.split(r'(?i)(Expression:\s*)', text)
    if len(parts) <= 1:
        return text

    result_parts = [parts[0]]  # text before first Expression:
    expression_count = 0

    i = 1
    while i < len(parts):
        tag   = parts[i]       # the "Expression: " token
        if i + 1 < len(parts):
            content = parts[i + 1]  # text that follows
        else:
            content = ''

        # Extract first clause (up to comma/semicolon/newline) for keyword check
        first_clause = re.split(r'[,;\n]', content)[0] if content else ''

        if expression_count < 2 and _is_facial(first_clause):
            result_parts.append(tag)   # keep Expression: prefix
            expression_count += 1
        # else: drop the "Expression:" prefix, keep the content below

        result_parts.append(content)
        i += 2

    result = ''.join(result_parts)
    # Clean up double commas or spaces left behind
    result = re.sub(r',\s*,', ',', result)
    result = re.sub(r'[ \t]{2,}', ' ', result)
    return result.strip()


def validate_prompt_count(prompts_list: list[dict], total_expected: int) -> dict:
    """Validate generated prompt count against expected total.

    Args:
        prompts_list: list of prompt dicts with 'block' key
        total_expected: total number of SRT subtitle blocks

    Returns:
        dict with keys: expected, generated, missing (list), extra (list), is_perfect (bool)
    """
    expected_set  = set(range(1, total_expected + 1))
    generated_set = set(p["block"] for p in prompts_list)

    missing = sorted(expected_set - generated_set)
    extra   = sorted(generated_set - expected_set)

    return {
        "expected":   total_expected,
        "generated":  len(generated_set),
        "missing":    missing,
        "extra":      extra,
        "is_perfect": len(missing) == 0 and len(extra) == 0,
    }


def export_txt(prompts: list[dict], mode: str) -> str:
    lines = []
    for p in prompts:
        image_prompt = clean_prompt_text(p['image_prompt'])
        lines.append(f"Image Prompt {p['block']}: {image_prompt}")
        if mode == "B" and p.get('video_prompt'):
            video_prompt = clean_prompt_text(p['video_prompt'])
            lines.append(f"Video Prompt {p['block']}: {video_prompt}")
        lines.append("")
    return '\n'.join(lines)


def export_xlsx(prompts: list[dict], mode: str, style_key: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Prompts"

    header_fill  = PatternFill('solid', fgColor='1a1a2e')
    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    thin_border  = Border(
        left=Side(style='thin', color='d0d0d0'),
        right=Side(style='thin', color='d0d0d0'),
        top=Side(style='thin', color='d0d0d0'),
        bottom=Side(style='thin', color='d0d0d0')
    )
    # History 2 Color/B&W row fills
    color_fill = PatternFill('solid', fgColor='FFF3E0')   # warm amber for Color rows
    bw_fill    = PatternFill('solid', fgColor='F5F5F5')   # light grey for B&W rows

    is_h2 = style_key == "history_2"
    is_h4 = style_key == "history_4"
    # History 4 word-count fills
    wc_ok_fill  = PatternFill('solid', fgColor='E8F5E9')   # light green — within target range
    wc_bad_fill = PatternFill('solid', fgColor='FFEBEE')   # light red   — outside target range

    if is_h2:
        if mode == "B":
            headers    = ['Block #', 'Type', 'Image Prompt', 'Video Prompt']
            col_widths = [8, 10, 80, 60]
        else:
            headers    = ['Block #', 'Type', 'Image Prompt']
            col_widths = [8, 10, 90]
    elif is_h4:
        if mode == "B":
            headers    = ['Block #', 'Words', 'Image Prompt', 'Video Prompt']
            col_widths = [8, 8, 80, 60]
        else:
            headers    = ['Block #', 'Words', 'Image Prompt']
            col_widths = [8, 8, 90]
    elif mode == "B":
        headers    = ['Block #', 'Image Prompt', 'Video Prompt']
        col_widths = [8, 80, 60]
    else:
        headers    = ['Block #', 'Image Prompt']
        col_widths = [8, 100]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64 + i)].width = w

    for row_idx, p in enumerate(prompts, 2):
        prompt_text = clean_prompt_text(p['image_prompt'])

        # ── Block # ──────────────────────────────────────────────────────────
        cell_num = ws.cell(row=row_idx, column=1, value=p['block'])
        cell_num.font      = Font(name='Arial', size=10, bold=True)
        cell_num.alignment = Alignment(horizontal='center', vertical='top')
        cell_num.border    = thin_border

        if is_h2:
            # ── Type column (Color / B&W) ─────────────────────────────────
            ptype      = _detect_prompt_type(prompt_text)
            row_fill   = bw_fill if ptype == "B&W" else color_fill
            type_label = "⬛ B&W" if ptype == "B&W" else "🎨 Color"

            cell_type            = ws.cell(row=row_idx, column=2, value=type_label)
            cell_type.font       = Font(name='Arial', size=10, bold=True)
            cell_type.alignment  = Alignment(horizontal='center', vertical='top')
            cell_type.border     = thin_border
            cell_type.fill       = row_fill

            cell_img             = ws.cell(row=row_idx, column=3, value=prompt_text)
            cell_img.font        = Font(name='Arial', size=10)
            cell_img.alignment   = Alignment(vertical='top', wrap_text=True)
            cell_img.border      = thin_border
            cell_img.fill        = row_fill

            if mode == "B":
                cell_vid           = ws.cell(row=row_idx, column=4, value=clean_prompt_text(p.get('video_prompt', '')))
                cell_vid.font      = Font(name='Arial', size=10)
                cell_vid.alignment = Alignment(vertical='top', wrap_text=True)
                cell_vid.border    = thin_border
                cell_vid.fill      = row_fill
        elif is_h4:
            # ── Word count column ─────────────────────────────────────────
            actual_wc = len(prompt_text.split())
            in_range  = 25 <= actual_wc <= 47   # ±3 of min/max targets
            wc_fill   = wc_ok_fill if in_range else wc_bad_fill

            cell_wc            = ws.cell(row=row_idx, column=2, value=actual_wc)
            cell_wc.font       = Font(name='Arial', size=10, bold=True)
            cell_wc.alignment  = Alignment(horizontal='center', vertical='top')
            cell_wc.border     = thin_border
            cell_wc.fill       = wc_fill

            cell_img           = ws.cell(row=row_idx, column=3, value=prompt_text)
            cell_img.font      = Font(name='Arial', size=10)
            cell_img.alignment = Alignment(vertical='top', wrap_text=True)
            cell_img.border    = thin_border
            cell_img.fill      = wc_fill

            if mode == "B":
                cell_vid           = ws.cell(row=row_idx, column=4, value=clean_prompt_text(p.get('video_prompt', '')))
                cell_vid.font      = Font(name='Arial', size=10)
                cell_vid.alignment = Alignment(vertical='top', wrap_text=True)
                cell_vid.border    = thin_border
        else:
            cell_img             = ws.cell(row=row_idx, column=2, value=prompt_text)
            cell_img.font        = Font(name='Arial', size=10)
            cell_img.alignment   = Alignment(vertical='top', wrap_text=True)
            cell_img.border      = thin_border

            if mode == "B":
                cell_vid           = ws.cell(row=row_idx, column=3, value=clean_prompt_text(p.get('video_prompt', '')))
                cell_vid.font      = Font(name='Arial', size=10)
                cell_vid.alignment = Alignment(vertical='top', wrap_text=True)
                cell_vid.border    = thin_border

        ws.row_dimensions[row_idx].height = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
